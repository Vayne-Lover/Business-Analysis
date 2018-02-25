# -*- coding: utf-8 -*-
# Author: vayne@zju.edu.cn

import time
import openpyxl as xl

def main():
    fileName = "movies.xlsx"

    wb = xl.Workbook()
    sheetName = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(sheetName[0])

    wb2 = xl.load_workbook(fileName)
    ori_sheetName = wb2.get_sheet_names()
    ori_ws = wb2.get_sheet_by_name(ori_sheetName[0])

    count = 1
    tempCount = 1
    while tempCount <= ori_ws.max_row:
        if count == 1:
            for i in range(0, 21):
                ws[chr(ord("A")+ i)+ str(count)].value = ori_ws[chr(ord("A")+ i)+ str(count)].value
            count += 1
            tempCount += 1
        elif ori_ws["O" + str(tempCount)].value is None:
            count += 1
            tempCount += 1
            # print(ori_ws["O" + str(count)].value)
            # print(count, tempCount)
        else:
            companies = ori_ws["O" + str(tempCount)].value.split("|")
            for j in range(0, len(companies)):
                for i in range(0, 21):
                    ws[chr(ord("A") + i) + str(count)].value = ori_ws[chr(ord("A") + i) + str(tempCount)].value
                ws["O" + str(count)].value = companies[j]
                count += 1
            tempCount += 1

    wb.save("q2.xlsx")

if __name__ == "__main__":
    start = time.clock()
    main()
    end = time.clock()
    print(end-start)
